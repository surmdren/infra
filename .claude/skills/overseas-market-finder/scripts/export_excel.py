#!/usr/bin/env python3
"""
Export market opportunities to Excel

Create formatted Excel spreadsheet with sortable/filterable opportunity list
"""

import json
from typing import List, Dict
from datetime import datetime


def export_to_excel(opportunities: List[Dict], output_file: str = None):
    """
    Export opportunities to Excel with formatting

    Args:
        opportunities: List of scored opportunities
        output_file: Output file path (default: auto-generated)

    Returns:
        Path to created Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
        return None

    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"market_opportunities_{timestamp}.xlsx"

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opportunities"

    # Define headers
    headers = [
        "Product", "Industry", "Segment", "Target Market",
        "Success Rate (%)", "Meets Threshold",
        "Market Size Score", "Competition Score",
        "Profitability Score", "Execution Score",
        "TAM (USD)", "Growth Rate (%)", "Competitors",
        "AOV (USD)", "Gross Margin (%)", "Payback (months)"
    ]

    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_num, opp in enumerate(opportunities, 2):
        scores = opp.get('scores', {})

        # Basic info
        ws.cell(row=row_num, column=1).value = opp.get('product', '')
        ws.cell(row=row_num, column=2).value = opp.get('industry', '')
        ws.cell(row=row_num, column=3).value = opp.get('segment', '')
        ws.cell(row=row_num, column=4).value = opp.get('target_market', '')

        # Success rate
        success_cell = ws.cell(row=row_num, column=5)
        success_rate = opp.get('success_rate', 0)
        success_cell.value = success_rate
        success_cell.number_format = '0.0'

        # Color code by success rate
        if success_rate >= 80:
            success_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            success_cell.font = Font(color="006100", bold=True)
        elif success_rate >= 60:
            success_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            success_cell.font = Font(color="9C5700")
        else:
            success_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            success_cell.font = Font(color="9C0006")

        # Meets threshold
        ws.cell(row=row_num, column=6).value = "✓" if opp.get('meets_threshold', False) else "✗"

        # Individual scores
        ws.cell(row=row_num, column=7).value = scores.get('market_size', 0)
        ws.cell(row=row_num, column=8).value = scores.get('competition', 0)
        ws.cell(row=row_num, column=9).value = scores.get('profitability', 0)
        ws.cell(row=row_num, column=10).value = scores.get('execution', 0)

        # Raw metrics
        ws.cell(row=row_num, column=11).value = opp.get('tam_usd', 0)
        ws.cell(row=row_num, column=11).number_format = '$#,##0'

        ws.cell(row=row_num, column=12).value = opp.get('growth_rate', 0) * 100
        ws.cell(row=row_num, column=12).number_format = '0.0'

        ws.cell(row=row_num, column=13).value = opp.get('num_competitors', 0)

        ws.cell(row=row_num, column=14).value = opp.get('avg_order_value', 0)
        ws.cell(row=row_num, column=14).number_format = '$#,##0'

        ws.cell(row=row_num, column=15).value = opp.get('gross_margin', 0) * 100
        ws.cell(row=row_num, column=15).number_format = '0.0'

        ws.cell(row=row_num, column=16).value = opp.get('payback_months', 0)

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = len(headers[col_num - 1])
        for row in range(2, len(opportunities) + 2):
            cell_value = str(ws.cell(row=row, column=col_num).value or '')
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Add filters
    ws.auto_filter.ref = ws.dimensions

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save workbook
    wb.save(output_file)
    print(f"✅ Excel file created: {output_file}")
    return output_file


def main():
    """Example usage"""
    # Example opportunities
    opportunities = [
        {
            'product': 'Smart Home Sensors',
            'industry': 'IoT',
            'segment': 'Home Automation',
            'target_market': 'North America',
            'success_rate': 85.5,
            'meets_threshold': True,
            'scores': {
                'market_size': 90,
                'competition': 75,
                'profitability': 88,
                'execution': 82
            },
            'tam_usd': 5000000000,
            'growth_rate': 0.25,
            'num_competitors': 15,
            'avg_order_value': 150,
            'gross_margin': 0.45,
            'payback_months': 8
        }
    ]

    export_to_excel(opportunities, 'market_opportunities_example.xlsx')


if __name__ == '__main__':
    main()
