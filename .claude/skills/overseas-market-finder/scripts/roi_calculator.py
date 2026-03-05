#!/usr/bin/env python3
"""
ROI Calculator for Overseas Market Opportunities

Calculate investment returns, break-even period, and financial projections
for market opportunities across conservative, normal, and optimistic scenarios.
"""

from typing import Dict, List
from datetime import datetime


class ROICalculator:
    """Calculate ROI and financial projections for market opportunities"""

    def __init__(self):
        self.scenarios = {
            'conservative': 0.5,  # 50% of normal
            'normal': 1.0,
            'optimistic': 1.5     # 150% of normal
        }

    def calculate_startup_costs(self, opportunity: Dict) -> Dict[str, float]:
        """
        Calculate startup costs breakdown

        Args:
            opportunity: Dict with product and market details

        Returns:
            Dict with cost categories
        """
        product_type = opportunity.get('product_type', 'physical')
        target_market = opportunity.get('target_market', 'North America')

        # Base costs (USD)
        costs = {
            'product_development': 0,
            'inventory': 0,
            'marketing': 0,
            'operations': 0,
            'certifications': 0,
            'total': 0
        }

        if product_type == 'physical':
            # Physical product costs
            costs['product_development'] = opportunity.get('mold_fee', 5000) + \
                                          opportunity.get('sample_cost', 2000)
            costs['inventory'] = opportunity.get('first_order_qty', 500) * \
                               opportunity.get('unit_cost', 20)
            costs['marketing'] = 10000  # Initial ads + listing
            costs['operations'] = 5000  # Platform fees, tools, etc.

            # Certification costs vary by market
            cert_costs = {
                'North America': 5000,  # UL, FCC
                'Europe': 8000,         # CE, ROHS
                'Southeast Asia': 3000  # Local certs
            }
            costs['certifications'] = cert_costs.get(target_market, 5000)

        elif product_type == 'saas':
            # SaaS product costs
            costs['product_development'] = opportunity.get('dev_cost', 30000)
            costs['inventory'] = 0  # No inventory
            costs['marketing'] = 15000  # SEO, ads, content
            costs['operations'] = 5000  # Hosting, tools
            costs['certifications'] = 2000  # GDPR compliance, security audits

        costs['total'] = sum(costs.values())
        return costs

    def calculate_monthly_revenue(self, opportunity: Dict, scenario: str = 'normal') -> Dict:
        """
        Calculate monthly revenue projections

        Args:
            opportunity: Market opportunity data
            scenario: 'conservative', 'normal', or 'optimistic'

        Returns:
            Dict with revenue breakdown
        """
        multiplier = self.scenarios[scenario]

        # Base projections
        avg_order_value = opportunity.get('avg_order_value', 100)
        gross_margin = opportunity.get('gross_margin', 0.4)
        monthly_sales = opportunity.get('projected_monthly_sales', 100)

        # Apply scenario multiplier
        adjusted_sales = monthly_sales * multiplier

        revenue = {
            'gross_revenue': adjusted_sales * avg_order_value,
            'cogs': adjusted_sales * avg_order_value * (1 - gross_margin),
            'gross_profit': adjusted_sales * avg_order_value * gross_margin,
            'units_sold': adjusted_sales
        }

        # Operating expenses (monthly)
        revenue['marketing_expense'] = revenue['gross_revenue'] * 0.20  # 20% of revenue
        revenue['operations_expense'] = 2000  # Fixed monthly ops
        revenue['net_profit'] = revenue['gross_profit'] - \
                               revenue['marketing_expense'] - \
                               revenue['operations_expense']

        return revenue

    def calculate_breakeven(self, startup_costs: float, monthly_net_profit: float) -> Dict:
        """
        Calculate break-even period

        Args:
            startup_costs: Total startup investment
            monthly_net_profit: Average monthly net profit

        Returns:
            Dict with break-even metrics
        """
        if monthly_net_profit <= 0:
            return {
                'breakeven_months': float('inf'),
                'breakeven_achievable': False,
                'message': 'Not profitable - cannot break even'
            }

        breakeven_months = startup_costs / monthly_net_profit

        return {
            'breakeven_months': round(breakeven_months, 1),
            'breakeven_achievable': breakeven_months <= 24,  # Within 2 years
            'message': f'Break-even in {round(breakeven_months, 1)} months' if breakeven_months <= 24
                      else 'Break-even period too long (>24 months)'
        }

    def calculate_roi_12m(self, startup_costs: float, net_profit_12m: float) -> float:
        """
        Calculate 12-month ROI

        Args:
            startup_costs: Initial investment
            net_profit_12m: Total net profit after 12 months

        Returns:
            ROI percentage
        """
        if startup_costs == 0:
            return 0

        roi = ((net_profit_12m - startup_costs) / startup_costs) * 100
        return round(roi, 1)

    def calculate_npv(self, startup_costs: float, monthly_cashflow: List[float],
                     discount_rate: float = 0.10) -> float:
        """
        Calculate Net Present Value (NPV) over 36 months

        Args:
            startup_costs: Initial investment (negative)
            monthly_cashflow: List of monthly cash flows
            discount_rate: Annual discount rate (default 10%)

        Returns:
            NPV in USD
        """
        monthly_rate = discount_rate / 12
        npv = -startup_costs

        for month, cashflow in enumerate(monthly_cashflow, start=1):
            npv += cashflow / ((1 + monthly_rate) ** month)

        return round(npv, 2)

    def analyze_opportunity(self, opportunity: Dict) -> Dict:
        """
        Complete financial analysis across all scenarios

        Args:
            opportunity: Complete opportunity data

        Returns:
            Comprehensive financial analysis
        """
        # Calculate startup costs
        startup_costs = self.calculate_startup_costs(opportunity)

        results = {
            'product': opportunity.get('product_name', 'Unknown'),
            'startup_costs': startup_costs,
            'scenarios': {}
        }

        # Analyze each scenario
        for scenario in ['conservative', 'normal', 'optimistic']:
            monthly_revenue = self.calculate_monthly_revenue(opportunity, scenario)

            # 12-month projection
            net_profit_12m = monthly_revenue['net_profit'] * 12

            # Break-even analysis
            breakeven = self.calculate_breakeven(
                startup_costs['total'],
                monthly_revenue['net_profit']
            )

            # ROI calculation
            roi_12m = self.calculate_roi_12m(startup_costs['total'], net_profit_12m)

            # NPV calculation (36 months)
            monthly_cashflow = [monthly_revenue['net_profit']] * 36
            npv_36m = self.calculate_npv(startup_costs['total'], monthly_cashflow)

            results['scenarios'][scenario] = {
                'monthly_revenue': monthly_revenue,
                'annual_net_profit': net_profit_12m,
                'breakeven': breakeven,
                'roi_12m': roi_12m,
                'npv_36m': npv_36m
            }

        # Overall recommendation
        normal_scenario = results['scenarios']['normal']
        results['recommendation'] = self._get_recommendation(
            normal_scenario['breakeven']['breakeven_months'],
            normal_scenario['roi_12m'],
            normal_scenario['npv_36m']
        )

        return results

    def _get_recommendation(self, breakeven_months: float, roi_12m: float,
                           npv_36m: float) -> str:
        """Generate investment recommendation"""
        if breakeven_months <= 12 and roi_12m >= 50 and npv_36m > 50000:
            return '⭐⭐⭐⭐⭐ 强烈推荐投资（快速回本+高回报）'
        elif breakeven_months <= 18 and roi_12m >= 30 and npv_36m > 20000:
            return '⭐⭐⭐⭐ 推荐投资（良好财务指标）'
        elif breakeven_months <= 24 and roi_12m >= 10 and npv_36m > 0:
            return '⭐⭐⭐ 可以考虑（需严格控制成本）'
        elif breakeven_months <= 36:
            return '⭐⭐ 谨慎投资（回本周期较长）'
        else:
            return '⭐ 不建议投资（财务风险高）'

    def generate_excel_report(self, analysis: Dict, output_file: str = None):
        """
        Generate Excel financial report

        Args:
            analysis: Financial analysis results
            output_file: Output Excel file path
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
            return None

        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"roi_analysis_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ROI Analysis"

        # Header
        ws['A1'] = f"ROI Analysis: {analysis['product']}"
        ws['A1'].font = Font(size=16, bold=True)

        # Startup Costs
        row = 3
        ws[f'A{row}'] = "Startup Costs"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        for category, amount in analysis['startup_costs'].items():
            ws[f'A{row}'] = category.replace('_', ' ').title()
            ws[f'B{row}'] = amount
            ws[f'B{row}'].number_format = '$#,##0'
            row += 1

        # Scenario Analysis
        row += 2
        ws[f'A{row}'] = "Scenario Analysis"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # Headers
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Conservative"
        ws[f'C{row}'] = "Normal"
        ws[f'D{row}'] = "Optimistic"
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        row += 1

        # Metrics
        metrics = [
            ('Monthly Net Profit', 'monthly_revenue', 'net_profit'),
            ('Annual Net Profit', 'annual_net_profit', None),
            ('Break-even (months)', 'breakeven', 'breakeven_months'),
            ('ROI 12M (%)', 'roi_12m', None),
            ('NPV 36M', 'npv_36m', None)
        ]

        for metric_name, key1, key2 in metrics:
            ws[f'A{row}'] = metric_name
            for col, scenario in enumerate(['conservative', 'normal', 'optimistic'], start=2):
                value = analysis['scenarios'][scenario][key1]
                if key2:
                    value = value[key2]

                cell = ws.cell(row=row, column=col)
                cell.value = value

                if 'Profit' in metric_name or 'NPV' in metric_name:
                    cell.number_format = '$#,##0'
                elif '%' in metric_name:
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = '0.0'
            row += 1

        # Recommendation
        row += 2
        ws[f'A{row}'] = "Recommendation"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = analysis['recommendation']
        ws.merge_cells(f'A{row}:D{row}')

        # Auto-adjust column widths
        for column in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[column].width = 20

        wb.save(output_file)
        print(f"✅ ROI analysis saved to: {output_file}")
        return output_file


def main():
    """Example usage"""
    calculator = ROICalculator()

    # Example opportunity
    opportunity = {
        'product_name': 'Smart Pet Feeder',
        'product_type': 'physical',
        'target_market': 'North America',
        'avg_order_value': 180,
        'gross_margin': 0.45,
        'projected_monthly_sales': 200,
        'mold_fee': 8000,
        'sample_cost': 2000,
        'first_order_qty': 500,
        'unit_cost': 35
    }

    analysis = calculator.analyze_opportunity(opportunity)

    # Print results
    print(f"\n{'='*60}")
    print(f"ROI Analysis: {analysis['product']}")
    print(f"{'='*60}\n")

    print("Startup Costs:")
    for category, amount in analysis['startup_costs'].items():
        print(f"  {category.replace('_', ' ').title()}: ${amount:,.0f}")

    print(f"\nScenario Analysis:")
    for scenario in ['conservative', 'normal', 'optimistic']:
        data = analysis['scenarios'][scenario]
        print(f"\n  {scenario.upper()}:")
        print(f"    Monthly Net Profit: ${data['monthly_revenue']['net_profit']:,.0f}")
        print(f"    Break-even: {data['breakeven']['breakeven_months']} months")
        print(f"    ROI (12M): {data['roi_12m']}%")
        print(f"    NPV (36M): ${data['npv_36m']:,.0f}")

    print(f"\n{analysis['recommendation']}\n")

    # Generate Excel
    calculator.generate_excel_report(analysis, 'roi_analysis_example.xlsx')


if __name__ == '__main__':
    main()
